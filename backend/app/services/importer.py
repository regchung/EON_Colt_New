"""車行批次訂單匯入:解析 xlsx / csv → 正規化 → 產生 OrderCreate 資料。

欄位以中文表頭辨識(支援多種別名),缺欄或格式錯誤會逐列回報,
不影響其他正確列的匯入。
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime, time, timedelta, timezone

TW = timezone(timedelta(hours=8))   # 車行「上車時間」為台灣本地時間,一律以 +08 儲存

from openpyxl import load_workbook

# 系統欄位 → 可接受的表頭別名(全部轉小寫、去空白後比對)
COLUMN_ALIASES: dict[str, list[str]] = {
    "service_date": ["服務日期", "日期", "預約日期", "service_date", "date"],
    "pickup_time": ["上車時間", "預約時間", "時間", "pickup_time", "time"],
    "pickup_window_min": ["彈性", "彈性分鐘", "彈性(分)", "時間彈性", "window"],
    "passenger_name": ["乘客", "乘客姓名", "姓名", "name", "passenger"],
    "passenger_phone": ["電話", "聯絡電話", "手機", "phone", "tel"],
    "pickup_address": ["上車地址", "上車地點", "起點", "pickup", "pickup_address"],
    "dropoff_address": ["下車地址", "下車地點", "迄點", "終點", "dropoff", "dropoff_address"],
    "pax": ["人數", "乘客數", "pax", "passengers"],
    "vehicle_type": ["車種", "車輛類型", "type", "vehicle_type"],
    "need_wheelchair": ["輪椅", "需要輪椅", "wheelchair", "need_wheelchair"],
    "allow_pool": ["共乘", "可共乘", "pool", "allow_pool"],
    "note": ["備註", "註記", "note", "remark"],
}

TRUE_TOKENS = {"y", "yes", "true", "1", "是", "v", "✓", "o", "需要"}
FALSE_TOKENS = {"n", "no", "false", "0", "否", "x", "不需要", ""}


def _norm(s) -> str:
    return str(s).strip().lower() if s is not None else ""


def _build_header_map(headers: list) -> dict[str, int]:
    """把實際表頭列對應到系統欄位 → 欄索引。"""
    alias_to_field = {}
    for field, aliases in COLUMN_ALIASES.items():
        for a in aliases:
            alias_to_field[_norm(a)] = field

    mapping: dict[str, int] = {}
    for idx, h in enumerate(headers):
        field = alias_to_field.get(_norm(h))
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip().replace("/", "-")
    return datetime.strptime(s, "%Y-%m-%d").date()


def _parse_time_to_dt(value, d: date) -> datetime:
    """把上車時間(可能是 'HH:MM'、datetime、time)組成當日 datetime(台灣 +08 tz-aware)。"""
    if isinstance(value, datetime):
        dt = value if value.date() != date(1900, 1, 1) else datetime.combine(d, value.time())
    elif isinstance(value, time):
        dt = datetime.combine(d, value)
    else:
        s = str(value).strip()
        t = datetime.strptime(s, "%H:%M").time() if ":" in s else time(0, 0)
        dt = datetime.combine(d, t)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=TW)
    return dt


def _parse_bool(value, default: bool) -> bool:
    s = _norm(value)
    if s in TRUE_TOKENS:
        return True
    if s in FALSE_TOKENS:
        return False
    return default


def _pool_consent(value) -> bool:
    """共乘同意規則:預設不同意,唯一例外為值含『同意』(且非『不同意』)。"""
    s = str(value).strip() if value is not None else ""
    return ("同意" in s) and ("不同意" not in s)


def _parse_vehicle_type(value) -> str:
    s = _norm(value)
    if "福祉" in s or "welfare" in s or "wheelchair" in s:
        return "welfare"
    return "normal"


def _read_rows(filename: str, content: bytes) -> tuple[list, list[list]]:
    """回傳 (表頭, 資料列)。支援 .xlsx 與 .csv。"""
    name = filename.lower()
    if name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    elif name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        rows = [r for r in csv.reader(io.StringIO(text))]
    else:
        raise ValueError("僅支援 .xlsx 或 .csv 檔")

    rows = [r for r in rows if any(c not in (None, "") for c in r)]  # 去空列
    if not rows:
        return [], []
    return rows[0], rows[1:]


def parse_orders(filename: str, content: bytes) -> tuple[list[dict], list[dict]]:
    """解析檔案。回傳 (可建立的訂單 payload list, 錯誤明細 list)。

    錯誤明細格式:{"row": <Excel 列號>, "error": "訊息"}
    """
    headers, data_rows = _read_rows(filename, content)
    if not headers:
        return [], [{"row": 0, "error": "檔案沒有內容"}]

    hmap = _build_header_map(headers)
    missing = [f for f in ("service_date", "pickup_address", "dropoff_address") if f not in hmap]
    if missing:
        labels = {"service_date": "服務日期", "pickup_address": "上車地址", "dropoff_address": "下車地址"}
        names = "、".join(labels[m] for m in missing)
        return [], [{"row": 1, "error": f"缺少必要欄位:{names}"}]

    def cell(row, field):
        idx = hmap.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    payloads: list[dict] = []
    errors: list[dict] = []

    for i, row in enumerate(data_rows):
        excel_row = i + 2  # 含表頭、1-based
        try:
            d = _parse_date(cell(row, "service_date"))
            pickup_dt = _parse_time_to_dt(cell(row, "pickup_time"), d)
            pickup_addr = str(cell(row, "pickup_address") or "").strip()
            dropoff_addr = str(cell(row, "dropoff_address") or "").strip()
            if not pickup_addr or not dropoff_addr:
                raise ValueError("上車/下車地址不可空白")

            vtype = _parse_vehicle_type(cell(row, "vehicle_type"))
            wheelchair = _parse_bool(cell(row, "need_wheelchair"), default=(vtype == "welfare"))

            payloads.append(
                {
                    "service_date": d.isoformat(),
                    "pickup_time": pickup_dt.isoformat(),
                    "pickup_window_min": int(cell(row, "pickup_window_min") or 30),
                    "passenger_name": (str(cell(row, "passenger_name")).strip() or None)
                    if cell(row, "passenger_name") else None,
                    "passenger_phone": (str(cell(row, "passenger_phone")).strip() or None)
                    if cell(row, "passenger_phone") else None,
                    "pickup_address": pickup_addr,
                    "dropoff_address": dropoff_addr,
                    "pax": int(cell(row, "pax") or 1),
                    "vehicle_type": vtype,
                    "need_wheelchair": wheelchair,
                    # 共乘規則:預設不同意,唯一例外為「共乘欄位值含『同意』」(否則需行控徵詢後才可併)
                    "allow_pool": _pool_consent(cell(row, "allow_pool")),
                    "note": (str(cell(row, "note")).strip() or None) if cell(row, "note") else None,
                    "status": "imported",
                }
            )
        except Exception as e:  # noqa: BLE001 逐列回報,不中斷整批
            errors.append({"row": excel_row, "error": str(e)})

    return payloads, errors
