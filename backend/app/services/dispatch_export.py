"""派遣表匯出(Excel)。

依日期 + 車行 產生:
- per_vehicle=False:單一檔多分頁(總覽 / 各子車隊 / 每車排班 / 派車明細 / 未派)。
- per_vehicle=True :每車一張工作表(派車單)+ 總覽。
回傳 .xlsx bytes。
"""
from __future__ import annotations

import io
import re
from collections import defaultdict
from datetime import date, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.order import Order
from app.models.unassigned_record import UnassignedRecord
from app.models.vehicle import Vehicle
from app.services import roster as roster_svc


def _unassigned_reason(o, ur: dict, has_welfare: bool) -> str:
    """未派原因:優先取對比引擎已推斷的 unassigned_record,否則依訂單屬性啟發式判斷。"""
    if o.id in ur and ur[o.id]:
        return ur[o.id]
    if o.pickup_lng is None or o.dropoff_lng is None:
        return "缺座標(地址無法地理編碼)"
    h = o.pickup_time.astimezone(TW).hour if o.pickup_time else 12
    if h < 6 or h > 18:
        return "服務時段外(上車不在 06:00–18:00)"
    if o.vehicle_type == "welfare":
        return "需福祉車:無福祉車可派或福祉車滿載" if not has_welfare else "福祉車滿載/時窗衝突,需增派或重排"
    return "車隊滿載或求解邊際(需增派/重排或放寬設定)"

TW = timezone(timedelta(hours=8))
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="2F5496")
_TITLE = Font(bold=True, size=14)
_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")


def _hm(dt) -> str:
    if not dt:
        return ""
    t = dt.astimezone(TW)
    return f"{t.hour:02d}:{t.minute:02d}"


def _style_header(ws, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws) -> None:
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(48, max(8, w * 1.6))


def _safe_sheet(name: str, used: set) -> str:
    n = _INVALID_SHEET.sub("-", name)[:31] or "車輛"
    base, i = n, 1
    while n in used:
        suffix = f"~{i}"
        n = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(n)
    return n


def _load(db: Session, service_date: date, fleet: str | None):
    q = select(Order).where(Order.service_date == service_date)
    if fleet:
        q = q.where(Order.fleet == fleet)
    orders = list(db.scalars(q).all())
    veh = {v.id: v for v in db.scalars(select(Vehicle)).all()}
    drv: dict[int, str] = {
        vid: info["name"]
        for vid, info in roster_svc.driver_for_date(db, service_date).items()
        if info.get("name")
    }
    return orders, veh, drv


def build_workbook(db: Session, service_date: date, fleet: str | None,
                   per_vehicle: bool, source: str = "auto") -> bytes:
    orders, veh, drv = _load(db, service_date, fleet)
    if source == "auto":
        # 以自動派遣落地(auto_dispatch_stop)覆蓋指派(in-memory,結束前 rollback 不落庫)
        from app.models.auto_dispatch_stop import AutoDispatchStop
        auto = {oid: (vid, seq, eta) for oid, vid, seq, eta in db.execute(
            select(AutoDispatchStop.order_id, AutoDispatchStop.vehicle_id,
                   AutoDispatchStop.seq, AutoDispatchStop.eta)
            .where(AutoDispatchStop.service_date == service_date,
                   AutoDispatchStop.kind == "pickup")).all()}
        for o in orders:
            a = auto.get(o.id)
            o.assigned_vehicle_id, o.dispatch_seq, o.eta = a if a else (None, None, None)
    asg = [o for o in orders if o.assigned_vehicle_id]
    un = [o for o in orders if not o.assigned_vehicle_id]
    byv: dict[int, list] = defaultdict(list)
    for o in asg:
        byv[o.assigned_vehicle_id].append(o)
    for vid in byv:
        if source == "auto":
            byv[vid].sort(key=lambda o: (o.eta or o.pickup_time))   # 自動:seq 跨車行會重複,依時間排
        else:
            byv[vid].sort(key=lambda o: (o.dispatch_seq or 0, o.pickup_time))

    wb = Workbook()
    scope = fleet or "全車行"

    # --- 總覽(兩種版型共用)---
    ws = wb.active
    ws.title = "總覽"
    ws.append([f"EON COLT 派遣表 — {service_date.isoformat()}（{scope}）"])
    ws["A1"].font = _TITLE
    ws.append([])
    ws.append(["指標", "數值"])
    for c in ws["3:3"]:
        c.font = Font(bold=True)
    used_v = len(byv)
    kpis = [
        ("訂單總數", len(orders)), ("已派", len(asg)), ("未派", len(un)),
        ("用車數", used_v),
        ("福祉車", sum(1 for v in byv if veh[v].type == "welfare")),
        ("一般車", sum(1 for v in byv if veh[v].type != "welfare")),
        ("趟數/車(平均)", round(len(asg) / used_v, 1) if used_v else 0),
        ("車行", scope),
    ]
    for k, v in kpis:
        ws.append([k, v])

    def _detail_row(v, o, seq):
        return [v.plate, drv.get(v.id, "-"), seq, _hm(o.eta or o.pickup_time),
                o.passenger_name, o.pax or 1,
                "福祉" if o.vehicle_type == "welfare" else "一般",
                o.pickup_address, o.dropoff_address, o.fleet,
                o.support_fleet if (o.support_fleet and o.support_fleet != o.fleet) else ""]

    DET_COLS = ["車牌", "司機", "順序", "上車ETA", "乘客", "人數", "車型", "上車地址", "下車地址",
                "子車隊", "支援車行"]

    if per_vehicle:
        # 每車一張工作表(派車單)
        used_names = {"總覽"}
        for vid, os in sorted(byv.items(), key=lambda kv: veh[kv[0]].plate):
            v = veh[vid]
            wsv = wb.create_sheet(_safe_sheet(v.plate, used_names))
            wsv.append([f"{v.plate}　{drv.get(vid, '-')}　{'福祉' if v.type == 'welfare' else '一般'}"
                        f"　{service_date.isoformat()}　{len(os)} 趟 / {sum(o.pax or 1 for o in os)} 人次"])
            wsv["A1"].font = _TITLE
            wsv.append(DET_COLS[2:])           # 單車表省略車牌/司機欄
            _style_header_row(wsv, 2, len(DET_COLS) - 2)
            for i, o in enumerate(os, 1):
                wsv.append(_detail_row(v, o, i)[2:])
            wsv.freeze_panes = "A3"
            _autofit(wsv)
    else:
        # 單一檔多分頁
        ws2 = wb.create_sheet("各子車隊")
        ws2.append(["子車隊", "訂單", "已派", "未派"])
        _style_header(ws2, 4)
        fl = defaultdict(lambda: [0, 0])
        for o in orders:
            fl[o.fleet or "(空)"][0] += 1
            if o.assigned_vehicle_id:
                fl[o.fleet or "(空)"][1] += 1
        for f, (t, a) in sorted(fl.items(), key=lambda x: -x[1][0]):
            ws2.append([f, t, a, t - a])

        ws3 = wb.create_sheet("每車排班")
        ws3.append(["車牌", "司機", "車型", "趟數", "人次", "首班", "末班"])
        _style_header(ws3, 7)
        for vid, os in sorted(byv.items(), key=lambda kv: -len(kv[1])):
            v = veh[vid]
            ts = sorted(o.pickup_time for o in os)
            ws3.append([v.plate, drv.get(vid, "-"), "福祉" if v.type == "welfare" else "一般",
                        len(os), sum(o.pax or 1 for o in os), _hm(ts[0]), _hm(ts[-1])])

        ws4 = wb.create_sheet("派車明細")
        ws4.append(DET_COLS)
        _style_header(ws4, len(DET_COLS))
        for vid, os in sorted(byv.items(), key=lambda kv: veh[kv[0]].plate):
            v = veh[vid]
            for i, o in enumerate(os, 1):
                ws4.append(_detail_row(v, o, i))
        ws4.freeze_panes = "A2"

    # 未派(兩種版型皆附)+ 未派原因
    ur = {r.order_id: r.reason_detail for r in db.scalars(
        select(UnassignedRecord).where(UnassignedRecord.service_date == service_date)).all()
        if r.order_id}
    has_welfare = bool(db.scalar(select(func.count()).select_from(Vehicle).where(
        Vehicle.type == "welfare", Vehicle.active.is_(True))) or 0)
    ws5 = wb.create_sheet("未派")
    ws5.append(["訂單ID", "乘客", "時間", "需求", "子車隊", "上車地址", "下車地址", "未派原因"])
    _style_header(ws5, 8)
    for o in un:
        ws5.append([o.id, o.passenger_name, _hm(o.pickup_time),
                    "福祉/輪椅" if o.vehicle_type == "welfare" else "一般",
                    o.fleet, o.pickup_address, o.dropoff_address,
                    _unassigned_reason(o, ur, has_welfare)])

    for w in wb.worksheets:
        _autofit(w)
    buf = io.BytesIO()
    wb.save(buf)
    if source == "auto":
        db.rollback()   # 丟棄 in-memory 指派覆蓋,不落庫
    return buf.getvalue()


def _style_header_row(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(horizontal="center")
