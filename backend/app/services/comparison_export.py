"""人工 vs 自動派遣比對匯出(Excel)。

- summary_workbook:某日期區間(可選車行)→ 逐日總覽 + 各車行日明細(取自 dispatch_comparison)。
- by_vehicle_workbook:某日某車行 → 逐車對比(同一組實體車牌,人工/自動並排)。
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.dispatch_comparison import DispatchComparison
from app.services import comparison as comparison_svc

_HEAD = Font(bold=True, color="FFFFFF")
_FILL = PatternFill("solid", fgColor="2F5496")
_TITLE = Font(bold=True, size=14)
_BOLD = Font(bold=True)


def _hdr(ws, ncol: int, row: int = 1) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.font = _HEAD
        cell.fill = _FILL
        cell.alignment = Alignment(horizontal="center")


def _fit(ws) -> None:
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(42, max(8, w * 1.5))


def _pct(h: int, v: int) -> float:
    return round(100 * (h - v) / h, 1) if h else 0.0


def summary_workbook(db: Session, date_from: date, date_to: date, fleet: str | None) -> bytes:
    q = select(DispatchComparison).where(
        DispatchComparison.service_date >= date_from,
        DispatchComparison.service_date <= date_to,
    )
    if fleet:
        q = q.where(DispatchComparison.fleet == fleet)
    rows = list(db.scalars(q.order_by(DispatchComparison.service_date, DispatchComparison.fleet)).all())

    wb = Workbook()
    scope = fleet or "全車行"

    # 逐日總覽
    ws = wb.active
    ws.title = "比對總覽"
    ws.append([f"人工 vs 自動派遣比對　{date_from.isoformat()} ～ {date_to.isoformat()}（{scope}）"])
    ws["A1"].font = _TITLE
    ws.append([])
    ws.append(["日期", "訂單", "人工用車", "自動用車", "省車", "省車率%", "未派"])
    _hdr(ws, 7, row=3)
    by_day: dict[date, list[int]] = defaultdict(lambda: [0, 0, 0, 0])  # orders,human,auto,unassigned
    for r in rows:
        d = by_day[r.service_date]
        d[0] += r.n_orders; d[1] += r.human_vehicles; d[2] += r.vroom_vehicles; d[3] += r.vroom_unassigned
    T = [0, 0, 0, 0]
    for d in sorted(by_day):
        o, h, v, u = by_day[d]
        ws.append([d.isoformat(), o, h, v, h - v, _pct(h, v), u])
        for i in range(4):
            T[i] += [o, h, v, u][i]
    ws.append(["合計", T[0], T[1], T[2], T[1] - T[2], _pct(T[1], T[2]), T[3]])
    for c in ws[ws.max_row]:
        c.font = _BOLD

    # 各車行日明細
    ws2 = wb.create_sheet("各車行日")
    ws2.append(["日期", "車行", "訂單", "人工用車", "自動用車", "省車", "省車率%", "未派",
                "人工里程km", "自動行駛分"])
    _hdr(ws2, 10)
    for r in rows:
        ws2.append([r.service_date.isoformat(), r.fleet, r.n_orders, r.human_vehicles,
                    r.vroom_vehicles, r.saved_vehicles, _pct(r.human_vehicles, r.vroom_vehicles),
                    r.vroom_unassigned, round((r.human_distance_m or 0) / 1000, 1),
                    round((r.vroom_drive_sec or 0) / 60)])
    ws2.freeze_panes = "A2"

    for w in wb.worksheets:
        _fit(w)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def by_vehicle_workbook(db: Session, fleet: str, service_date: date, window_min: int | None = None) -> bytes:
    r = comparison_svc.compare_day_by_vehicle(db, fleet, service_date, window_min)
    wb = Workbook()
    ws = wb.active
    ws.title = "逐車對比"
    if not r:
        ws.append([f"{service_date.isoformat()} {fleet}:無可對比資料(需成行單 + 人工派遣紀錄)"])
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    t = r["totals"]
    ws.append([f"{service_date.isoformat()}　{fleet}　逐車對比"
               f"（人工 {t['human']['vehicles']} 車 vs 自動 {t['auto']['vehicles']} 車）"])
    ws["A1"].font = _TITLE
    ws.append(["車牌", "駕駛", "車型", "座位", "人工趟數", "自動趟數", "人工出車", "自動出車"])
    _hdr(ws, 8, row=2)
    for v in r["vehicles"]:
        ws.append([v["plate"], v.get("driver") or "-",
                   "福祉" if v["type"] == "welfare" else "一般", v.get("seats"),
                   len(v.get("human") or []), len(v.get("auto") or []),
                   "是" if v.get("human_used") else "", "是" if v.get("auto_used") else ""])
    ws.append([])
    ws.append(["合計", "", "", "",
               f"{t['human']['vehicles']}車 / {t['human']['distance_km']}km / {t['human']['work_min']}分",
               f"{t['auto']['vehicles']}車 / {t['auto']['distance_km']}km / {t['auto']['work_min']}分", "", ""])
    for c in ws[ws.max_row]:
        c.font = _BOLD
    ws.freeze_panes = "A3"
    _fit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
